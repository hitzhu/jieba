import openpyxl
# 创建一个工作簿
wb = openpyxl.Workbook()

# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook('微信群聊天记录.xlsx')
# 第二步：选取表单
sh = wb['sheet1']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row = 1,column = 1)   # 读取第一行，第一列的数据
#print(ce.value)
# 按行读取数据 list(sh.rows)
#print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
for cases in list(sh.rows):

    case_data = cases[3].value
    print(case_data)
# 关闭工作薄
wb.close()